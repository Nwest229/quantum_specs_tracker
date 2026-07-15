"""Excel workbook exporter for the quantum backend tracker.

Produces a multi-sheet .xlsx that mirrors the JSON contract:

  * "Tracker"      one row per backend, every column from the spec, with the
                   Tech-source and Price-source cells rendered as clickable
                   hyperlinks and their retrieval dates in their own columns.
  * "ChartData"    numeric, cross-sectional helper table (price + specs +
                   derived $/QV, $/AQ) that the native charts read from.
  * "Charts"       native Excel charts: Per-shot price vs #Qubits (scatter) and
                   $ per Quantum Volume by backend (bar).

openpyxl is required (``pip install openpyxl``); it is optional for the rest of
the pipeline. Values keep the scraper's honesty: blanks stay blank, never zero.
"""
from __future__ import annotations

import os
import re
from typing import Any, Callable, Optional

UNKNOWN = "Not publicly disclosed"
_NUM_RE = re.compile(r"[-+]?\d[\d,]*\.?\d*(?:[eE][-+]?\d+)?")


def _known(v: Any) -> bool:
    return v is not None and v != UNKNOWN and v != ""


def _num(v: Any) -> Optional[float]:
    if not _known(v):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    m = _NUM_RE.search(str(v))
    return float(m.group().replace(",", "")) if m else None


def _parse_price(text: Any) -> Optional[dict]:
    """'USD 0.08' -> {value:0.08, currency:'USD'} ; None if no number."""
    if not _known(text):
        return None
    s = str(text)
    m = _NUM_RE.search(s)
    if not m:
        return None
    cur = "USD"
    for tok, code in (("EUR", "EUR"), ("€", "EUR"), ("GBP", "GBP"), ("£", "GBP"),
                      ("USD", "USD"), ("$", "USD")):
        if tok in s:
            cur = code
            break
    return {"value": float(m.group().replace(",", "")), "currency": cur}


def _leaf(doc: dict, group: str, key: str) -> Any:
    v = (doc.get(group, {}) or {}).get(key)
    return v if _known(v) else None


def _prov(doc: dict, name: str) -> Any:
    v = doc.get(name)
    if isinstance(v, dict):
        v = v.get("value")
    return v if _known(v) else None


def _source(doc: dict, *prefixes: str) -> tuple[Optional[str], Optional[str]]:
    """First (url, retrieved) among sources whose field starts with any prefix."""
    for s in doc.get("sources", []):
        f = str(s.get("field", ""))
        if any(f.startswith(p) for p in prefixes):
            return s.get("url"), s.get("retrieved")
    return None, None


def _tech_source(doc: dict) -> tuple[Optional[str], Optional[str]]:
    # first non-pricing source = the technical provenance
    for s in doc.get("sources", []):
        if not str(s.get("field", "")).startswith("pricing."):
            return s.get("url"), s.get("retrieved")
    return None, None


def _theo_log2(doc: dict) -> Any:
    tm = (doc.get("derived_metrics", {}) or {}).get("theoretical_max", {})
    return _num(tm.get("log2_value")) if isinstance(tm, dict) else None


# (header, accessor) — accessor returns a plain cell value (str/number/None).
_COLUMNS: list[tuple[str, Callable[[dict], Any]]] = [
    ("ID", lambda d: d.get("id")),
    ("Type", lambda d: d.get("type") if _known(d.get("type")) else None),
    ("Vendor", lambda d: d.get("vendor")),
    ("Model", lambda d: d.get("model") if _known(d.get("model")) else None),
    ("System Name", lambda d: d.get("system_name") if _known(d.get("system_name")) else None),
    ("Backend Name", lambda d: d.get("backend_name")),
    ("Planned Release", lambda d: d.get("planned_release") if _known(d.get("planned_release")) else None),
    ("Commercial Release", lambda d: d.get("commercial_release") if _known(d.get("commercial_release")) else None),
    ("Argmax", lambda d: _prov(d, "argmax")),
    ("Black-box (AQ)", lambda d: _prov(d, "black_box")),
    ("Vendor Metric", lambda d: _prov(d, "vendor_metric")),
    ("Theo. Max 2^ (log2)", _theo_log2),
    ("Quantum Volume", lambda d: _prov(d, "quantum_volume")),
    ("#Qubits", lambda d: _leaf(d, "qpu_topology", "qubits")),
    ("Topology", lambda d: _leaf(d, "qpu_topology", "type")),
    ("#Edges", lambda d: _leaf(d, "qpu_topology", "edges")),
    ("2Q max", lambda d: _leaf(d, "fidelity", "2q_max")),
    ("2Q avg", lambda d: _leaf(d, "fidelity", "2q_avg")),
    ("2Q median", lambda d: _leaf(d, "fidelity", "2q_median")),
    ("2Q min", lambda d: _leaf(d, "fidelity", "2q_min")),
    ("1Q max", lambda d: _leaf(d, "fidelity", "1q_max")),
    ("1Q avg", lambda d: _leaf(d, "fidelity", "1q_avg")),
    ("1Q min", lambda d: _leaf(d, "fidelity", "1q_min")),
    ("SPAM avg", lambda d: _leaf(d, "fidelity", "spam_avg")),
    ("1Q gate time (s)", lambda d: _leaf(d, "operation_speed", "1q_gate_time_s")),
    ("2Q gate time (s)", lambda d: _leaf(d, "operation_speed", "2q_gate_time_s")),
    ("Readout time (s)", lambda d: _leaf(d, "operation_speed", "readout_time_s")),
    ("Shot rate min (Hz)", lambda d: _leaf(d, "operation_speed", "shot_rate_min")),
    ("Shot rate avg (Hz)", lambda d: _leaf(d, "operation_speed", "shot_rate_avg")),
    ("Shot rate max (Hz)", lambda d: _leaf(d, "operation_speed", "shot_rate_max")),
    ("CLOPS (Hz)", lambda d: _leaf(d, "operation_speed", "clops")),
    ("Credits/Hour", lambda d: _leaf(d, "operation_speed", "credits_per_hour")),
    ("Mid-circuit Meas.", lambda d: _leaf(d, "features", "mid_circuit_measurement")),
    ("MCM conditional", lambda d: _leaf(d, "features", "conditional_logic")),
    ("Parallel 2Q", lambda d: _leaf(d, "features", "parallel_2q")),
    ("Qubit reuse", lambda d: _leaf(d, "features", "qubit_reuse")),
    ("Hybrid", lambda d: _leaf(d, "features", "hybrid_execution")),
    ("Uptime", lambda d: _leaf(d, "features", "uptime")),
    ("Per-1Q gate", lambda d: _leaf(d, "pricing", "per_1q_gate")),
    ("Per-2Q gate", lambda d: _leaf(d, "pricing", "per_2q_gate")),
    ("Per-iteration", lambda d: _leaf(d, "pricing", "per_iteration")),
    ("Per-shot", lambda d: _leaf(d, "pricing", "per_shot")),
    ("Per-task", lambda d: _leaf(d, "pricing", "per_task")),
    ("Price/second", lambda d: _leaf(d, "pricing", "per_second")),
    ("Price/hour", lambda d: _leaf(d, "pricing", "per_hour")),
    ("Price/month", lambda d: _leaf(d, "pricing", "per_month")),
    ("Price/System", lambda d: _leaf(d, "pricing", "per_system")),
    ("Comment", lambda d: _leaf(d, "pricing", "comments")),
]

# Columns holding provenance, inserted next to their groups.
_TECH_SRC_AFTER = "Credits/Hour"        # tech source + date go after operation speed
_PRICE_SRC_AFTER = "Comment"            # price source + date go at the end


def _require_openpyxl():
    try:
        import openpyxl  # noqa: F401
        return openpyxl
    except ImportError as e:  # pragma: no cover
        raise ImportError(
            "openpyxl is required for --xlsx export. Install it with:\n"
            "    pip install openpyxl\n"
            "(or `pip install -r requirements.txt`)."
        ) from e


def _price_usd(doc_price: Any) -> Optional[float]:
    p = _parse_price(doc_price)
    if p and p.get("currency") == "USD":
        return p["value"]
    return None


def write_workbook(docs: list[dict], out_path: str) -> str:
    openpyxl = _require_openpyxl()
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import ScatterChart, BarChart, Reference, Series

    wb = openpyxl.Workbook()
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="374151")
    link_font = Font(color="2563EB", underline="single")

    # ---- Tracker sheet ---------------------------------------------------
    ws = wb.active
    ws.title = "Tracker"

    # Build the header list with the two source/date column pairs interleaved.
    headers: list[str] = []
    for h, _ in _COLUMNS:
        headers.append(h)
        if h == _TECH_SRC_AFTER:
            headers += ["Tech Date retrieved", "Tech Source"]
        if h == _PRICE_SRC_AFTER:
            headers += ["Price Date retrieved", "Price Source"]

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "G2"  # freeze through Backend Name + header row
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    for r, doc in enumerate(docs, start=2):
        tech_url, tech_ret = _tech_source(doc)
        price_url, price_ret = _source(doc, "pricing.")
        col = 1
        for h, fn in _COLUMNS:
            ws.cell(row=r, column=col, value=fn(doc))
            col += 1
            if h == _TECH_SRC_AFTER:
                ws.cell(row=r, column=col, value=tech_ret); col += 1
                _write_link(ws, r, col, tech_url, link_font); col += 1
            if h == _PRICE_SRC_AFTER:
                ws.cell(row=r, column=col, value=price_ret); col += 1
                _write_link(ws, r, col, price_url, link_font); col += 1

    _autosize(ws, headers, get_column_letter)

    # ---- ChartData sheet (numeric, cross-sectional) ----------------------
    cd = wb.create_sheet("ChartData")
    cd_headers = ["Backend", "Vendor", "#Qubits", "Per-shot USD", "Per-task USD",
                  "Per-second USD", "Quantum Volume", "Algorithmic Qubits",
                  "2Q avg fidelity", "Theo max log2", "USD/QV", "USD/AQ"]
    for c, h in enumerate(cd_headers, 1):
        cell = cd.cell(row=1, column=c, value=h)
        cell.font = head_font
        cell.fill = head_fill
    cd_rows = 0
    for doc in docs:
        pshot = _price_usd(_leaf(doc, "pricing", "per_shot"))
        ptask = _price_usd(_leaf(doc, "pricing", "per_task"))
        psec = _price_usd(_leaf(doc, "pricing", "per_second"))
        qv = _num(_prov(doc, "quantum_volume"))
        aq = _num(_prov(doc, "black_box"))
        qb = _num(_leaf(doc, "qpu_topology", "qubits"))
        fid = _num(_leaf(doc, "fidelity", "2q_avg"))
        tm = _theo_log2(doc)
        # only keep rows that carry at least a price and a spec (chartable)
        if pshot is None and psec is None:
            continue
        headline = pshot if pshot is not None else psec
        row = [doc.get("backend_name"), doc.get("vendor"), qb, pshot, ptask, psec,
               qv, aq, fid, tm,
               (headline / qv) if (qv and headline is not None) else None,
               (headline / aq) if (aq and headline is not None) else None]
        cd_rows += 1
        for c, v in enumerate(row, 1):
            cd.cell(row=cd_rows + 1, column=c, value=v)
    _autosize(cd, cd_headers, get_column_letter)

    # ---- Charts sheet (cross-sectional) ---------------------------------
    ch = wb.create_sheet("Charts")
    if cd_rows >= 2:
        # Scatter: Per-shot price (Y) vs #Qubits (X)
        sc = ScatterChart()
        sc.title = "Per-shot price vs #Qubits"
        sc.x_axis.title = "#Qubits"
        sc.y_axis.title = "Per-shot price (USD)"
        sc.style = 13
        xref = Reference(cd, min_col=3, min_row=2, max_row=cd_rows + 1)
        yref = Reference(cd, min_col=4, min_row=1, max_row=cd_rows + 1)
        series = Series(yref, xref, title_from_data=True)
        series.marker.symbol = "circle"
        series.graphicalProperties.line.noFill = True
        sc.series.append(series)
        ch.add_chart(sc, "A1")

        # Bar: USD/QV by backend
        bc = BarChart()
        bc.title = "USD per Quantum Volume (lower = cheaper capability)"
        bc.type = "col"
        bc.y_axis.title = "USD / QV"
        data = Reference(cd, min_col=11, min_row=1, max_row=cd_rows + 1)
        cats = Reference(cd, min_col=1, min_row=2, max_row=cd_rows + 1)
        bc.add_data(data, titles_from_data=True)
        bc.set_categories(cats)
        ch.add_chart(bc, "A18")

    os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)
    wb.save(out_path)
    return out_path


def _write_link(ws, row, col, url, link_font) -> None:
    cell = ws.cell(row=row, column=col)
    if url:
        cell.value = url
        cell.hyperlink = url
        cell.font = link_font


def _autosize(ws, headers, get_column_letter, cap=48) -> None:
    for c in range(1, len(headers) + 1):
        letter = get_column_letter(c)
        longest = len(str(headers[c - 1]))
        for cell in ws[letter][1:]:
            if cell.value is not None:
                longest = max(longest, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(cap, max(10, longest + 2))
